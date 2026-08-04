"""Файл 1 «СС.2026.Таблицы» — выручка + транспорт Олег + дебиторка.
Берём из листа «Сводная» (левый блок: объект×месяц) и «Задолженность».
Выручка (метод согласован, сходится с контролем 130 312 854 за янв–июн):
  выручка = «Счёт клиенту» − трансфер ЗетТЕК (трансфер ЗетТЕК — наш расход, не выручка).
Транспорт Олег (переменный расход) = трансфер ЗетТЕК."""
import datetime, re
from collections import defaultdict
from . import config as C
from .util import load, header_index, num

RU = {"янв":1,"фев":2,"мар":3,"апр":4,"май":5,"июн":6,
      "июл":7,"авг":8,"сен":9,"окт":10,"ноя":11,"дек":12}


def _ym(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.year:04d}-{v.month:02d}"
    s = str(v).strip().lower()
    for k, n in RU.items():
        if s.startswith(k):
            mo = re.search(r"(\d{2})$", s)
            yr = "20" + mo.group(1) if mo else "2026"
            return f"{yr}-{n:02d}"
    return None


def _is_zettek(obj):
    return "зеттек" in str(obj).lower()


def _row_ip(row, iIP, iBank):
    """ИП строки-счёта из файла: сперва столбец «ИП», затем банк с суффиксом Д/П
    (Д = Миронов, П = Молчанов). Вернёт 'Миронов' | 'Молчанов' | None."""
    if iIP is not None and iIP < len(row) and row[iIP]:
        s = str(row[iIP]).lower()
        if "мирон" in s: return "Миронов"
        if "молчан" in s: return "Молчанов"
    if iBank is not None and iBank < len(row) and row[iBank]:
        b = str(row[iBank]).strip()
        if b.endswith("Д"): return "Миронов"
        if b.endswith("П"): return "Молчанов"
    return None


class IncomeData:
    def __init__(self):
        self.rev_obj_month = defaultdict(lambda: defaultdict(float))   # obj -> month -> выручка
        self.transport_month = defaultdict(float)                      # транспорт Олег (всего)
        self.transport_obj_month = defaultdict(lambda: defaultdict(float))  # транспорт по объектам
        self.hours_obj_month = defaultdict(lambda: defaultdict(float))    # obj -> month -> часы
        self.workcost_obj_month = defaultdict(lambda: defaultdict(float)) # obj -> month -> стоимость работ (расчёт ставка×часы)
        # история ставок: obj -> смена -> month -> [Σ(ставка клиента×часы), Σ(ставка рабоч×часы), Σчасы]
        self.rates = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0.0, 0.0, 0.0])))
        self.receivables = []   # [{obj, оплачено, оказано, дебиторка, ндс}]
        self.years_seen = set()   # какие годы вообще есть в листах (сторож смены года)
        self.client_objects = []  # имена всех распознанных листов-клиентов (для авто-детекта новых)

    def revenue(self, frm, to, obj=None, ip=None):
        total = 0.0
        for o, months in self.rev_obj_month.items():
            if obj and o != obj:
                continue
            if ip and ip != "Все" and C.ip_of(o) != ip:
                continue
            for m, v in months.items():
                if frm <= m <= to:
                    total += v
        return total

    def revenue_by_month(self, frm, to, obj=None, ip=None):
        out = defaultdict(float)
        for o, months in self.rev_obj_month.items():
            if obj and o != obj:
                continue
            if ip and ip != "Все" and C.ip_of(o) != ip:
                continue
            for m, v in months.items():
                if frm <= m <= to:
                    out[m] += v
        return out

    def transport(self, frm, to):
        return sum(v for m, v in self.transport_month.items() if frm <= m <= to)

    def objects(self):
        return sorted(self.rev_obj_month.keys())


def parse_income(path=None):
    """Всё из ПЕРВИЧКИ — клиентских листов (по клеткам сверяется с таблицей Анны):
      • Выручка без НДС = столбец «Сумма услуг», строки где «Дата услуг ОТ» = дата (строку «итого» отсекаем), месяц по этой дате.
      • Транспорт Олег = столбец «Трансфер» дневного блока, месяц по столбцу «Дата».
      • Дебиторка = столбец «Недоплата по услугам». НДС = «Из них НДС». Счёт с НДС = «Сумма счёта с НДС»."""
    path = path or C.INCOME_XLSX
    import openpyxl, warnings
    warnings.simplefilter("ignore")
    wb = openpyxl.load_workbook(path, data_only=True)
    d = IncomeData()

    for nm in wb.sheetnames:
        ws = wb[nm]
        H = {str(c.value).strip().lower(): j for j, c in enumerate(ws[1]) if c.value}
        # лист-клиент опознаём по СИГНАТУРЕ (столбцы выручки), а не по «(26)» —
        # так авто-подхватываются новые клиенты и смена года в названиях листов.
        is_client = any(k == "сумма услуг" for k in H) and any("дата услуг от" in k for k in H)
        if any(s in nm.lower() for s in C.NON_CLIENT_SHEETS):
            is_client = False
        if not is_client:
            continue
        obj = nm.strip()
        def col(pred):
            for k, j in H.items():
                if pred(k):
                    return j
            return None

        # --- блок счетов: выручка (Сумма услуг) + дебиторка (Недоплата) + ИП ---
        iL = col(lambda k: "дата услуг от" in k)
        iN = col(lambda k: k == "сумма услуг")
        iSch = col(lambda k: "счет" in k and "ндс" in k)
        iVat = col(lambda k: "из них ндс" in k)
        iX = col(lambda k: "недоплата" in k)
        iIP = col(lambda k: k == "ип")                       # явный столбец «ИП»
        iBank = col(lambda k: "банк" in k)                   # банк с суффиксом Д/П
        ip_sum = defaultdict(float)                          # выручка объекта по ИП (для определения ИП объекта)
        okaz = sch = vat = debt = 0.0
        if iL is not None and iN is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not isinstance(row[iL], (datetime.datetime, datetime.date)):
                    continue   # строка «итого» = текст -> отсекаем
                # сторож смены года считает только строки С ДЕНЬГАМИ: в шаблонах полно
                # пустых заготовок с датами вперёд (04.01.2027) и мусорных 01.01.1900
                if num(row[iN]) or (iX is not None and num(row[iX])):
                    d.years_seen.add(row[iL].year)
                if row[iL].year != C.YEAR:
                    continue   # ЛИСТЫ ПРОШЛЫХ ЛЕТ. Архив «… (25) стар» имеет ту же сигнатуру
                               # столбцов, и без этой строки его старый долг (87 млн) втекал
                               # в дебиторку 2026-го. Выручку спасал фильтр по месяцам, её нет.
                m = f"{row[iL].year:04d}-{row[iL].month:02d}"
                n = num(row[iN])
                d.rev_obj_month[obj][m] += n          # ВЫРУЧКА = «Сумма услуг» по месяцу оказания
                okaz += n
                if iSch is not None: sch += num(row[iSch])
                if iVat is not None: vat += num(row[iVat])
                if iX is not None:   debt += num(row[iX])
                rip = _row_ip(row, iIP, iBank)        # ИП строки-счёта из файла
                if rip and n:
                    ip_sum[rip] += n
            # ИП объекта = ИП, на который выставлено больше выручки (из файла, не по имени)
            if ip_sum:
                C.OBJ_IP[obj] = max(ip_sum, key=ip_sum.get)
            if okaz or debt:
                d.client_objects.append(obj)   # клиент отчётного года (архивные листы сюда не попадают)
                d.receivables.append({"obj": obj, "оказано": okaz, "счет_ндс": sch,
                                      "ндс": vat, "долг": debt})

        # --- дневной блок: транспорт Олег + часы + стоимость работ (месяц по «Дата») ---
        iDate = col(lambda k: k == "дата")
        iTr = col(lambda k: "трансфер" in k)
        iHours = col(lambda k: "сумма часов" in k) or col(lambda k: "всего отработано" in k)
        iWork = col(lambda k: "итого за рабоч" in k) or col(lambda k: "стоимость работ" in k)
        iSm = col(lambda k: "смена" in k)
        iCr = col(lambda k: "час" in k and ("стоим" in k or "ст-ть" in k or "цена" in k))
        iWr = col(lambda k: "ставка" in k and "час" in k)
        if iDate is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                m = _ym(row[iDate])
                if not m:
                    continue
                g = lambda i: row[i] if (i is not None and i < len(row) and isinstance(row[i], (int, float))) else 0
                if iTr is not None and g(iTr):
                    d.transport_month[m] += g(iTr)
                    d.transport_obj_month[obj][m] += g(iTr)
                hrs = g(iHours)
                if hrs:
                    d.hours_obj_month[obj][m] += hrs
                if iWork is not None and g(iWork):
                    d.workcost_obj_month[obj][m] += g(iWork)
                # история ставок по смене (ставки могут быть строкой с запятой: '331,80')
                if hrs:
                    def numr(i):
                        if i is None or i >= len(row):
                            return 0.0
                        v = row[i]
                        if isinstance(v, (int, float)):
                            return float(v)
                        try:
                            return float(str(v).replace(" ", "").replace("\xa0", "").replace(",", "."))
                        except (ValueError, TypeError):
                            return 0.0
                    sm = str(row[iSm]).strip() if (iSm is not None and row[iSm]) else "—"
                    cell = d.rates[obj][sm][m]
                    cell[0] += numr(iCr) * hrs
                    cell[1] += numr(iWr) * hrs
                    cell[2] += hrs

    return d


def accrued_rev(inc, obj, month):
    """Начисленная выручка объекта за месяц = Σ(ставка клиента × часы) из дневного блока."""
    return sum(c[month][0] for c in inc.rates.get(obj, {}).values() if month in c)
